"""
Excel generator per guru.

Inputs (read):
  data/snapshots/holdings-{guru}-{quarter}.json
  data/snapshots/diff-{guru}-{from_q}-{to_q}.json  (optional)

Output:
  outputs/excel/{guru}-{quarter}.xlsx
  Sheets: "Current Holdings", "Quarterly Changes", "Top Movers"

Usage:
  python3 generate.py                       # all gurus, latest quarter
  python3 generate.py --guru berkshire_hathaway
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print(
        "error: openpyxl not installed. run: pip install -r requirements.txt",
        file=sys.stderr,
    )
    sys.exit(2)

ROOT = Path(__file__).resolve().parents[2]
SNAPSHOTS = ROOT / "data" / "snapshots"
OUT_DIR = Path(__file__).resolve().parent

HEADER_FILL = PatternFill(start_color="14171F", end_color="14171F", fill_type="solid")
HEADER_FONT = Font(name="Heebo", size=12, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Heebo", size=11)
POS_FONT = Font(name="Heebo", size=11, color="0F8A3C", bold=True)
NEG_FONT = Font(name="Heebo", size=11, color="C43030", bold=True)


def load_json(p: Path) -> dict:
    with p.open("r", encoding="utf-8") as f:
        return json.load(f)


def style_header(ws: Worksheet, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="right", vertical="center")


def autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_current_holdings(wb: Workbook, holdings: dict) -> None:
    ws = wb.create_sheet("Current Holdings")
    ws.sheet_view.rightToLeft = True
    headers = ["טיקר", "חברה", "מניות", "שווי (USD)", "אחוז מתיק", "שינוי %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))
    positions = sorted(
        holdings.get("positions", []),
        key=lambda p: p.get("value_usd", 0),
        reverse=True,
    )
    for idx, p in enumerate(positions, start=2):
        ws.cell(row=idx, column=1, value=p.get("ticker"))
        ws.cell(row=idx, column=2, value=p.get("company"))
        ws.cell(row=idx, column=3, value=p.get("shares"))
        ws.cell(row=idx, column=4, value=p.get("value_usd"))
        ws.cell(row=idx, column=5, value=p.get("pct_of_portfolio"))
        cell = ws.cell(row=idx, column=6, value=p.get("change_pct"))
        chg = p.get("change_pct") or 0
        cell.font = POS_FONT if chg > 0 else NEG_FONT if chg < 0 else BODY_FONT
        ws.cell(row=idx, column=3).number_format = "#,##0"
        ws.cell(row=idx, column=4).number_format = "#,##0"
        ws.cell(row=idx, column=5).number_format = "0.0\"%\""
        cell.number_format = "+0.0\"%\";-0.0\"%\";0.0\"%\""
    autosize(ws, [12, 30, 14, 18, 12, 12])
    ws.freeze_panes = "A2"


def sheet_quarterly_changes(wb: Workbook, diff: dict | None) -> None:
    ws = wb.create_sheet("Quarterly Changes")
    ws.sheet_view.rightToLeft = True
    if not diff:
        ws["A1"] = "אין נתוני diff זמינים לרבעון הזה"
        ws["A1"].font = BODY_FONT
        autosize(ws, [60])
        return

    headers = ["סוג שינוי", "טיקר", "מניות (לפני)", "מניות (אחרי)", "שווי (USD)", "שינוי %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))
    row = 2

    def write_row(ch_type: str, ticker: str, before, after, value, delta) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=ch_type)
        ws.cell(row=row, column=2, value=ticker)
        ws.cell(row=row, column=3, value=before)
        ws.cell(row=row, column=4, value=after)
        ws.cell(row=row, column=5, value=value)
        cell = ws.cell(row=row, column=6, value=delta)
        if isinstance(delta, (int, float)):
            cell.font = POS_FONT if delta > 0 else NEG_FONT if delta < 0 else BODY_FONT
            cell.number_format = "+0.0\"%\";-0.0\"%\";0.0\"%\""
        for col in (3, 4, 5):
            c = ws.cell(row=row, column=col)
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0"
        row += 1

    for n in diff.get("new", []):
        write_row("פוזיציה חדשה", n.get("ticker"), 0, n.get("shares"), n.get("value_usd"), None)
    for ex in diff.get("exited", []):
        write_row("יציאה", ex.get("ticker"), ex.get("shares_sold"), 0, ex.get("value_usd"), None)
    for inc in diff.get("increased", []):
        write_row(
            "הגדלה",
            inc.get("ticker"),
            inc.get("old_shares"),
            inc.get("new_shares"),
            None,
            inc.get("delta_pct"),
        )
    for dec in diff.get("decreased", []):
        write_row(
            "הקטנה",
            dec.get("ticker"),
            dec.get("old_shares"),
            dec.get("new_shares"),
            None,
            dec.get("delta_pct"),
        )

    autosize(ws, [16, 12, 16, 16, 18, 12])
    ws.freeze_panes = "A2"


def sheet_top_movers(wb: Workbook, diff: dict | None) -> None:
    ws = wb.create_sheet("Top Movers")
    ws.sheet_view.rightToLeft = True
    if not diff:
        ws["A1"] = "אין נתוני diff זמינים לרבעון הזה"
        ws["A1"].font = BODY_FONT
        autosize(ws, [60])
        return

    items: list[tuple[str, str, float]] = []
    for inc in diff.get("increased", []):
        items.append((inc.get("ticker"), "הגדלה", inc.get("delta_pct", 0)))
    for dec in diff.get("decreased", []):
        items.append((dec.get("ticker"), "הקטנה", dec.get("delta_pct", 0)))
    for n in diff.get("new", []):
        items.append((n.get("ticker"), "חדש", 100.0))
    for ex in diff.get("exited", []):
        items.append((ex.get("ticker"), "יציאה", -100.0))

    items.sort(key=lambda x: abs(x[2]), reverse=True)
    items = items[:20]

    headers = ["דירוג", "טיקר", "סוג", "שינוי %"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))
    for rank, (tk, kind, delta) in enumerate(items, start=1):
        ws.cell(row=rank + 1, column=1, value=rank)
        ws.cell(row=rank + 1, column=2, value=tk)
        ws.cell(row=rank + 1, column=3, value=kind)
        cell = ws.cell(row=rank + 1, column=4, value=delta)
        cell.font = POS_FONT if delta > 0 else NEG_FONT
        cell.number_format = "+0.0\"%\";-0.0\"%\";0.0\"%\""
    autosize(ws, [8, 12, 16, 12])
    ws.freeze_panes = "A2"


def find_latest_diff_for(guru: str, quarter: str) -> Path | None:
    pattern = re.compile(rf"^diff-{re.escape(guru)}-.*-{re.escape(quarter)}\.json$")
    candidates = [p for p in SNAPSHOTS.glob("diff-*.json") if pattern.match(p.name)]
    if not candidates:
        return None
    return sorted(candidates)[-1]


def generate_for(holdings_path: Path) -> Path:
    holdings = load_json(holdings_path)
    guru = holdings["guru"]
    quarter = holdings["quarter"]
    diff_path = find_latest_diff_for(guru, quarter)
    diff = load_json(diff_path) if diff_path else None

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    sheet_current_holdings(wb, holdings)
    sheet_quarterly_changes(wb, diff)
    sheet_top_movers(wb, diff)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / f"{guru}-{quarter}.xlsx"
    tmp = out.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    tmp.replace(out)
    return out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--guru", help="guru id (e.g. berkshire_hathaway)")
    args = parser.parse_args()

    if not SNAPSHOTS.exists():
        print(f"error: snapshots dir not found: {SNAPSHOTS}", file=sys.stderr)
        return 1

    pattern = "holdings-*.json"
    if args.guru:
        pattern = f"holdings-{args.guru}-*.json"

    files = sorted(SNAPSHOTS.glob(pattern))
    if not files:
        print(f"warn: no holdings files matched {pattern}", file=sys.stderr)
        return 1

    # group by guru, take latest quarter for each
    by_guru: dict[str, Path] = {}
    for f in files:
        # holdings-{guru}-{quarter}.json
        try:
            data = load_json(f)
            guru = data["guru"]
        except (json.JSONDecodeError, KeyError, OSError):
            continue
        prev = by_guru.get(guru)
        if prev is None or load_json(f)["quarter"] > load_json(prev)["quarter"]:
            by_guru[guru] = f

    for guru, f in by_guru.items():
        try:
            out = generate_for(f)
            print(f"ok: {out}")
        except Exception as e:
            print(f"error generating for {guru}: {e}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
